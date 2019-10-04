from openpyxl import load_workbook
import requests
import json


load_wb = load_workbook("address.xlsx", data_only=True)
# 시트 이름으로 불러오기 시트 스펠링 대소문자 주의.
load_ws = load_wb['sheet1']
api_key = "여기에 구글 geocoder API키만 입력하세요."
# 셀 주소로 값 출력
#print(load_ws['A1'].value)

# 셀 좌표로 값 출력
#print(load_ws.cell(1, 1).value)
## 데이터 가져오기

rowCount = 1
for row in load_ws.rows:
	#예외처리.
	try:
		address = load_ws.cell(rowCount, 1).value +" "+ load_ws.cell(rowCount, 2).value +" "+ load_ws.cell(rowCount, 3).value +" "+ load_ws.cell(rowCount, 4).value
		print(address)
	except:
		break
	request = requests.get(
		"https://maps.googleapis.com/maps/api/geocode/json?address=" + address + "%20110&language=ko&sensor=false&key=" + api_key)
	data = request.text
	search_result = json.loads(data)

	geolist = []
	#위경도값을 가져온다.
	for j in range(len(search_result['results'])):
		geolist.append(search_result['results'][j]['geometry']['location']['lat'])
		geolist.append(search_result['results'][j]['geometry']['location']['lng'])


	print(geolist)

		## cell 설정 [ B1 ~ B* : 위도 / C1 ~ C* : 경도]
	lat_cell = load_ws.cell(row=rowCount, column=5)
	lng_cell = load_ws.cell(row=rowCount, column=6)
	lat_cell.value = geolist[0]
	lng_cell.value = geolist[1]
	rowCount += 1


## 데이터 저장
load_wb.save("address_to.xlsx")
